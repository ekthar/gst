from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook
import xlrd

from gst_hsn_tool.config import LEARNING_FUZZY_THRESHOLD
from gst_hsn_tool.utils import normalize_hsn_digits, normalize_text


@dataclass
class LearnedRecord:
    input_description_norm: str
    input_category_norm: str
    input_client_hsn_norm: str
    resolved_hsn8: str
    resolved_description: str
    resolved_category: str
    resolved_rate: str
    match_type: str
    score: float
    learned_date: str
    use_count: int


class LearningMemory:
    FIELDNAMES = [
        "input_description_norm",
        "input_category_norm",
        "input_client_hsn_norm",
        "resolved_hsn8",
        "resolved_description",
        "resolved_category",
        "resolved_rate",
        "match_type",
        "score",
        "learned_date",
        "use_count",
    ]

    def __init__(self, csv_path: Path) -> None:
        self.csv_path = csv_path
        self.records: Dict[tuple[str, str, str], LearnedRecord] = {}
        self.desc_index: Dict[str, List[LearnedRecord]] = {}
        self.load()

    def load(self) -> None:
        self.records = {}
        self.desc_index = {}

        if not self.csv_path.exists():
            return

        with self.csv_path.open("r", newline="", encoding="utf-8-sig") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                desc = str(row.get("input_description_norm", "")).strip()
                if not desc:
                    continue
                category = str(row.get("input_category_norm", "")).strip()
                client_hsn = str(row.get("input_client_hsn_norm", "")).strip()
                resolved_hsn8 = str(row.get("resolved_hsn8", "")).strip()
                if not resolved_hsn8:
                    continue

                try:
                    score = float(row.get("score", 0) or 0)
                except ValueError:
                    score = 0.0
                try:
                    use_count = int(row.get("use_count", 0) or 0)
                except ValueError:
                    use_count = 0

                record = LearnedRecord(
                    input_description_norm=desc,
                    input_category_norm=category,
                    input_client_hsn_norm=client_hsn,
                    resolved_hsn8=resolved_hsn8,
                    resolved_description=str(row.get("resolved_description", "")),
                    resolved_category=str(row.get("resolved_category", "")),
                    resolved_rate=str(row.get("resolved_rate", "")),
                    match_type=str(row.get("match_type", "")),
                    score=score,
                    learned_date=str(row.get("learned_date", "")),
                    use_count=use_count,
                )

                key = (record.input_description_norm, record.input_category_norm, record.input_client_hsn_norm)
                existing = self.records.get(key)
                if existing is None or record.use_count >= existing.use_count:
                    self.records[key] = record

        for record in self.records.values():
            self.desc_index.setdefault(record.input_description_norm, []).append(record)

    def lookup_exact(self, description_norm: str, category_norm: str, client_hsn_norm: str) -> LearnedRecord | None:
        composite_key = (description_norm, category_norm, client_hsn_norm)
        if composite_key in self.records:
            return self.records[composite_key]

        broad_key = (description_norm, category_norm, "")
        if broad_key in self.records:
            return self.records[broad_key]

        candidates = self.desc_index.get(description_norm, [])
        if not candidates:
            return None

        if category_norm:
            category_hits = [c for c in candidates if c.input_category_norm == category_norm]
            if category_hits:
                return sorted(category_hits, key=lambda r: r.use_count, reverse=True)[0]

        return sorted(candidates, key=lambda r: r.use_count, reverse=True)[0]

    def lookup_fuzzy(self, description_norm: str, category_norm: str) -> LearnedRecord | None:
        if not description_norm:
            return None

        best: tuple[float, LearnedRecord] | None = None
        next_best = 0.0

        for desc_key, records in self.desc_index.items():
            similarity = SequenceMatcher(None, description_norm, desc_key).ratio() * 100
            if similarity < LEARNING_FUZZY_THRESHOLD:
                continue

            candidate = sorted(records, key=lambda r: (r.use_count, r.score), reverse=True)[0]
            weighted = similarity + min(5.0, candidate.use_count * 0.2)
            if category_norm and candidate.input_category_norm == category_norm:
                weighted += 2.0

            if best is None or weighted > best[0]:
                if best is not None:
                    next_best = max(next_best, best[0])
                best = (weighted, candidate)
            else:
                next_best = max(next_best, weighted)

        if best is None:
            return None

        if (best[0] - next_best) < 3.0:
            return None

        return best[1]

    def upsert_many(self, entries: List[dict]) -> int:
        if not entries:
            return 0

        now = datetime.now().isoformat(timespec="seconds")
        updated = 0

        for entry in entries:
            desc = str(entry.get("input_description_norm", "")).strip()
            if not desc:
                continue
            hsn8 = str(entry.get("resolved_hsn8", "")).strip()
            if not hsn8:
                continue

            key = (
                desc,
                str(entry.get("input_category_norm", "")).strip(),
                str(entry.get("input_client_hsn_norm", "")).strip(),
            )

            existing = self.records.get(key)
            if existing is None:
                record = LearnedRecord(
                    input_description_norm=key[0],
                    input_category_norm=key[1],
                    input_client_hsn_norm=key[2],
                    resolved_hsn8=hsn8,
                    resolved_description=str(entry.get("resolved_description", "")),
                    resolved_category=str(entry.get("resolved_category", "")),
                    resolved_rate=str(entry.get("resolved_rate", "")),
                    match_type=str(entry.get("match_type", "")),
                    score=float(entry.get("score", 0) or 0),
                    learned_date=now,
                    use_count=1,
                )
                self.records[key] = record
                self.desc_index.setdefault(record.input_description_norm, []).append(record)
                updated += 1
                continue

            existing.resolved_hsn8 = hsn8
            existing.resolved_description = str(entry.get("resolved_description", existing.resolved_description))
            existing.resolved_category = str(entry.get("resolved_category", existing.resolved_category))
            existing.resolved_rate = str(entry.get("resolved_rate", existing.resolved_rate))
            existing.match_type = str(entry.get("match_type", existing.match_type))
            try:
                incoming_score = float(entry.get("score", existing.score) or existing.score)
            except ValueError:
                incoming_score = existing.score
            existing.score = incoming_score
            existing.learned_date = now
            existing.use_count += 1
            updated += 1

        if updated:
            self._save_all()
        return updated

    def _save_all(self) -> None:
        self.csv_path.parent.mkdir(parents=True, exist_ok=True)
        with self.csv_path.open("w", newline="", encoding="utf-8") as handle:
            writer = csv.DictWriter(handle, fieldnames=self.FIELDNAMES)
            writer.writeheader()
            for _, record in sorted(self.records.items(), key=lambda item: item[0]):
                writer.writerow(
                    {
                        "input_description_norm": record.input_description_norm,
                        "input_category_norm": record.input_category_norm,
                        "input_client_hsn_norm": record.input_client_hsn_norm,
                        "resolved_hsn8": record.resolved_hsn8,
                        "resolved_description": record.resolved_description,
                        "resolved_category": record.resolved_category,
                        "resolved_rate": record.resolved_rate,
                        "match_type": record.match_type,
                        "score": record.score,
                        "learned_date": record.learned_date,
                        "use_count": record.use_count,
                    }
                )


def import_learning_file(
    file_path: Path,
    memory_csv_path: Path,
    product_header: str,
    category_header: str,
    hsn_header: str,
) -> dict:
    rows = _read_table_rows(file_path)
    if not rows:
        return {
            "file": str(file_path),
            "total_rows": 0,
            "usable_rows": 0,
            "saved_rows": 0,
            "skipped_rows": 0,
        }

    entries = []
    usable = 0
    skipped = 0

    p_header = product_header.strip().lower()
    c_header = category_header.strip().lower()
    h_header = hsn_header.strip().lower()

    for row in rows:
        key_map = {str(k).strip().lower(): row.get(k, "") for k in row.keys()}
        product = str(key_map.get(p_header, "")).strip()
        category = str(key_map.get(c_header, "")).strip()
        hsn_raw = str(key_map.get(h_header, "")).strip()

        if not product:
            skipped += 1
            continue

        hsn = normalize_hsn_digits(hsn_raw)
        if len(hsn) not in {4, 6, 8}:
            skipped += 1
            continue

        usable += 1
        entries.append(
            {
                "input_description_norm": normalize_text(product),
                "input_category_norm": normalize_text(category),
                "input_client_hsn_norm": "",
                "resolved_hsn8": hsn,
                "resolved_description": product,
                "resolved_category": normalize_text(category),
                "resolved_rate": "",
                "match_type": "manual_training_import",
                "score": 95,
            }
        )

    memory = LearningMemory(memory_csv_path)
    saved = memory.upsert_many(entries)

    return {
        "file": str(file_path),
        "total_rows": len(rows),
        "usable_rows": usable,
        "saved_rows": saved,
        "skipped_rows": skipped,
    }


def _read_table_rows(path: Path) -> List[dict]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return _read_csv(path)
    if suffix == ".xlsx":
        return _read_excel(path)
    if suffix == ".xls":
        return _read_xls(path)
    raise ValueError(f"Unsupported file type: {path}")


def _read_csv(path: Path) -> List[dict]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        return [dict(row) for row in reader]


def _read_excel(path: Path) -> List[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    values = list(ws.iter_rows(values_only=True))
    if not values:
        wb.close()
        return []

    headers = ["" if h is None else str(h).strip() for h in values[0]]
    out = []
    for row in values[1:]:
        obj = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            value = row[i] if i < len(row) else ""
            obj[header] = "" if value is None else str(value)
        out.append(obj)

    wb.close()
    return out


def _read_xls(path: Path) -> List[dict]:
    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_index(0)
    if sheet.nrows == 0:
        return []

    headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
    out = []
    for r in range(1, sheet.nrows):
        obj = {}
        for c, header in enumerate(headers):
            if not header:
                continue
            value = sheet.cell_value(r, c)
            obj[header] = "" if value is None else str(value)
        out.append(obj)
    return out
