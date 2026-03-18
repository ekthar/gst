from __future__ import annotations

import csv
import tempfile
import urllib.request
from pathlib import Path
from typing import Iterable, List, Tuple

from openpyxl import load_workbook

from gst_hsn_tool.utils import normalize_hsn_digits

GST_DIRECTORY_URL = "https://tutorial.gst.gov.in/downloads/HSN_SAC.xlsx"


def download_official_directory(destination_xlsx: Path) -> Path:
    destination_xlsx.parent.mkdir(parents=True, exist_ok=True)
    urllib.request.urlretrieve(GST_DIRECTORY_URL, destination_xlsx)
    return destination_xlsx


def transform_hsn_rows(rows: Iterable[Tuple[object, object]]) -> List[dict]:
    out = []
    seen = set()

    for code_value, description_value in rows:
        code = normalize_hsn_digits(code_value)
        if len(code) != 8:
            continue
        if code in seen:
            continue

        description = "" if description_value is None else str(description_value).strip()
        if not description:
            continue

        seen.add(code)
        out.append(
            {
                "hsn8": code,
                "description": description,
                "category": f"chapter_{code[:2]}",
                "rate": "",
                "aliases": "",
                "source": "gst_official_hsn_mstr",
            }
        )

    return out


def build_master_from_official_xlsx(input_xlsx: Path, output_csv: Path) -> int:
    wb = load_workbook(input_xlsx, read_only=True, data_only=True)
    try:
        if "HSN_MSTR" not in wb.sheetnames:
            raise ValueError("HSN_MSTR sheet not found in official GST directory file.")

        ws = wb["HSN_MSTR"]
        rows = ws.iter_rows(min_row=2, values_only=True)
        master_rows = transform_hsn_rows(rows)
        if not master_rows:
            raise ValueError("No valid 8-digit HSN entries found in HSN_MSTR.")
    finally:
        wb.close()

    output_csv.parent.mkdir(parents=True, exist_ok=True)
    with output_csv.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["hsn8", "description", "category", "rate", "aliases", "source"],
        )
        writer.writeheader()
        writer.writerows(master_rows)

    return len(master_rows)


def download_and_build_master(output_csv: Path) -> int:
    with tempfile.TemporaryDirectory(prefix="gst_hsn_") as temp_dir:
        raw_path = Path(temp_dir) / "HSN_SAC.xlsx"
        download_official_directory(raw_path)
        return build_master_from_official_xlsx(raw_path, output_csv)
