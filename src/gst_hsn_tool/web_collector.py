from __future__ import annotations

import csv
import hashlib
import html
import json
import re
import time
import urllib.parse
import urllib.request
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait
from collections import deque
from pathlib import Path
from typing import Callable, Iterable, List, Set, Tuple

from openpyxl import load_workbook

from gst_hsn_tool.utils import normalize_hsn_digits, normalize_text

DEFAULT_SEED_URLS = [
    "https://services.gst.gov.in/services/searchhsnsac",
    "https://tutorial.gst.gov.in/downloads/HSN_SAC.xlsx",
    "https://cleartax.in/s/gst-hsn-lookup",
]

USER_AGENT = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) GST-HSN-Collector/1.0"
GOOGLE_SEARCH_URL = "https://www.google.com/search"

DEFAULT_GOOGLE_QUERIES = [
    "gst hsn code list india",
    "cbic hsn tariff schedule",
    "hsn code chapter wise goods",
    "gst rate hsn code pdf",
    "india customs tariff hsn",
]

ASSET_EXTENSIONS = {
    ".css",
    ".js",
    ".ico",
    ".png",
    ".jpg",
    ".jpeg",
    ".svg",
    ".gif",
    ".webp",
    ".woff",
    ".woff2",
    ".ttf",
    ".eot",
}

RELEVANT_TOKENS = {
    "gst",
    "hsn",
    "sac",
    "tariff",
    "rate",
    "rates",
    "goods",
    "services",
    "customs",
    "code",
    "dataset",
    "download",
    "trade",
    "exim",
}

EXCLUDE_TOKENS = {
    "about",
    "about-us",
    "contact",
    "faq",
    "privacy",
    "policy",
    "terms",
    "career",
    "vacancy",
    "internship",
    "sitemap",
    "gallery",
    "event",
    "newsroom",
    "feedback",
}


def collect_online_hsn_pairs(
    raw_dir: Path,
    output_csv: Path,
    seed_urls: List[str],
    max_pages: int = 80,
    max_pages_per_domain: int = 120,
    max_seconds: int = 600,
    logger: Callable[[str], None] | None = None,
) -> dict:
    raw_dir.mkdir(parents=True, exist_ok=True)
    output_csv.parent.mkdir(parents=True, exist_ok=True)

    queue = deque(_normalize_url(url) for url in seed_urls if url.strip())
    visited: Set[str] = set()
    queued: Set[str] = set(queue)
    domain_counts: dict[str, int] = {}
    harvested_rows = []
    downloaded_files = []
    filtered_links = 0
    started = time.time()
    timed_out = False

    while queue and len(visited) < max_pages:
        if (time.time() - started) >= max_seconds:
            timed_out = True
            if logger:
                logger(f"Crawl time budget reached ({max_seconds}s). Stopping safely.")
            break

        url = queue.popleft().strip()
        if not url or url in visited:
            continue

        domain = urllib.parse.urlparse(url).netloc.lower()
        if domain_counts.get(domain, 0) >= max_pages_per_domain:
            continue

        visited.add(url)
        domain_counts[domain] = domain_counts.get(domain, 0) + 1
        if logger:
            logger(f"Collecting page {len(visited)}/{max_pages}: {url}")

        try:
            body, content_type = _fetch_url(url)
        except Exception:
            continue

        parsed = urllib.parse.urlparse(url)
        ext = Path(parsed.path).suffix.lower()

        if _is_download_like(content_type, ext):
            file_path = raw_dir / _safe_name_from_url(url)
            try:
                file_path.write_bytes(body)
            except Exception:
                continue
            downloaded_files.append(str(file_path))
            harvested_rows.extend(_extract_pairs_from_file(file_path))
            if logger:
                logger(f"Downloaded file source: {file_path}")
            continue

        text = _decode_text(body)
        harvested_rows.extend(_extract_pairs_from_text(text, source=url))

        for link in _extract_links(url, text):
            if link in visited:
                continue
            if not _allow_domain(seed_urls, link):
                continue
            if not _is_relevant_link(link):
                filtered_links += 1
                continue
            norm_link = _normalize_url(link)
            if norm_link in visited or norm_link in queued:
                continue
            queue.append(norm_link)
            queued.add(norm_link)

    unique = _dedup_pairs(harvested_rows)
    _write_pairs_csv(output_csv, unique)
    if logger:
        logger(f"Saved harvested HSN pairs: {output_csv} ({len(unique)} rows)")

    return {
        "pages_visited": len(visited),
        "filtered_links": filtered_links,
        "timed_out": timed_out,
        "downloaded_files": downloaded_files,
        "pair_count": len(unique),
        "pairs_file": str(output_csv),
    }


def collect_google_search_hsn_pairs(
    raw_dir: Path,
    output_csv: Path,
    queries: List[str] | None = None,
    max_results_per_query: int = 25,
    max_pages: int = 250,
    max_seconds: int = 600,
    max_workers: int = 8,
    discovery_request_delay: float = 0.35,
    logger: Callable[[str], None] | None = None,
) -> dict:
    raw_dir.mkdir(parents=True, exist_ok=True)
    output_csv.parent.mkdir(parents=True, exist_ok=True)

    active_queries = [q.strip() for q in (queries or DEFAULT_GOOGLE_QUERIES) if q.strip()]
    if not active_queries:
        active_queries = DEFAULT_GOOGLE_QUERIES[:]

    discovered_urls = _discover_urls_from_google(
        active_queries,
        max_results_per_query=max_results_per_query,
        request_delay_seconds=discovery_request_delay,
        logger=logger,
    )

    discovered_file = raw_dir / f"google_discovered_{int(time.time())}.csv"
    _write_discovered_urls_csv(discovered_file, discovered_urls)

    harvested_rows = []
    downloaded_files = []
    pages_visited = 0
    filtered_links = 0
    started = time.time()
    timed_out = False

    relevant_urls = []
    for url in discovered_urls:
        if _is_relevant_link(url):
            relevant_urls.append(url)
        else:
            filtered_links += 1

    candidate_urls = relevant_urls[:max_pages]
    worker_count = max(1, int(max_workers))
    in_flight = {}
    index = 0

    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        while index < len(candidate_urls) or in_flight:
            if (time.time() - started) >= max_seconds:
                timed_out = True
                if logger:
                    logger(f"Crawl time budget reached ({max_seconds}s). Stopping safely.")
                break

            while index < len(candidate_urls) and len(in_flight) < worker_count:
                url = candidate_urls[index]
                index += 1
                pages_visited += 1
                if logger:
                    logger(f"Collecting Google result {pages_visited}/{max_pages}: {url}")
                future = executor.submit(_fetch_and_extract_url, url, raw_dir)
                in_flight[future] = url

            if not in_flight:
                continue

            done, _ = wait(list(in_flight.keys()), timeout=0.5, return_when=FIRST_COMPLETED)
            for future in done:
                url = in_flight.pop(future)
                try:
                    result = future.result()
                except Exception:
                    continue

                harvested_rows.extend(result.get("rows", []))
                downloaded_file = str(result.get("downloaded_file", "")).strip()
                if downloaded_file:
                    downloaded_files.append(downloaded_file)

        if timed_out:
            for future in in_flight:
                future.cancel()

    unique = _dedup_pairs(harvested_rows)
    _write_pairs_csv(output_csv, unique)
    if logger:
        logger(f"Saved harvested HSN pairs: {output_csv} ({len(unique)} rows)")

    return {
        "queries_used": len(active_queries),
        "google_urls_discovered": len(discovered_urls),
        "google_discovered_file": str(discovered_file),
        "pages_visited": pages_visited,
        "filtered_links": filtered_links,
        "timed_out": timed_out,
        "downloaded_files": downloaded_files,
        "pair_count": len(unique),
        "pairs_file": str(output_csv),
    }


def _discover_urls_from_google(
    queries: List[str],
    max_results_per_query: int,
    request_delay_seconds: float = 0.35,
    logger: Callable[[str], None] | None = None,
) -> List[str]:
    out: List[str] = []
    seen: Set[str] = set()

    for query in queries:
        if logger:
            logger(f"Google search query: {query}")

        for start in range(0, max_results_per_query, 10):
            search_url = (
                f"{GOOGLE_SEARCH_URL}?q={urllib.parse.quote_plus(query)}"
                f"&num=10&hl=en&start={start}"
            )
            try:
                body, _ = _fetch_url(search_url)
            except Exception:
                break

            html_text = _decode_text(body)
            links = _extract_google_result_links(html_text)
            if not links:
                break

            for link in links:
                norm = _normalize_url(link)
                if norm in seen:
                    continue
                seen.add(norm)
                out.append(norm)

            if len(links) < 5:
                break

            if request_delay_seconds > 0:
                time.sleep(request_delay_seconds)

    return out


def _extract_google_result_links(html_text: str) -> List[str]:
    links: List[str] = []
    seen: Set[str] = set()

    # Standard Google result target format.
    for match in re.findall(r'href=["\']/url\?q=([^"\'&]+)', html_text, flags=re.IGNORECASE):
        url = urllib.parse.unquote(match)
        if not url.startswith(("http://", "https://")):
            continue
        if "google." in urllib.parse.urlparse(url).netloc.lower():
            continue
        if url not in seen:
            seen.add(url)
            links.append(url)

    # Fallback for additional direct absolute links in result block.
    for match in re.findall(r'href=["\'](https?://[^"\']+)["\']', html_text, flags=re.IGNORECASE):
        url = html.unescape(match)
        domain = urllib.parse.urlparse(url).netloc.lower()
        if "google." in domain:
            continue
        if url not in seen:
            seen.add(url)
            links.append(url)

    return links


def _write_discovered_urls_csv(path: Path, urls: List[str]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["url"])
        writer.writeheader()
        for url in urls:
            writer.writerow({"url": url})


def _fetch_and_extract_url(url: str, raw_dir: Path) -> dict:
    try:
        body, content_type = _fetch_url(url)
    except Exception:
        return {"rows": [], "downloaded_file": ""}

    parsed = urllib.parse.urlparse(url)
    ext = Path(parsed.path).suffix.lower()

    if _is_download_like(content_type, ext):
        file_path = raw_dir / _safe_name_from_url(url)
        if file_path.exists():
            digest = hashlib.md5(url.encode("utf-8", errors="ignore")).hexdigest()[:8]
            alt_name = f"{file_path.stem}_{digest}{file_path.suffix}"
            file_path = file_path.with_name(alt_name)
        try:
            file_path.write_bytes(body)
        except Exception:
            return {"rows": [], "downloaded_file": ""}
        return {
            "rows": _extract_pairs_from_file(file_path),
            "downloaded_file": str(file_path),
        }

    text = _decode_text(body)
    return {
        "rows": _extract_pairs_from_text(text, source=url),
        "downloaded_file": "",
    }


def _fetch_url(url: str) -> tuple[bytes, str]:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=20) as resp:
        content_type = str(resp.headers.get("Content-Type", "")).lower()
        body = resp.read()
    return body, content_type


def _decode_text(body: bytes) -> str:
    for encoding in ("utf-8", "utf-8-sig", "latin-1"):
        try:
            return body.decode(encoding, errors="ignore")
        except Exception:
            continue
    return body.decode(errors="ignore")


def _is_download_like(content_type: str, ext: str) -> bool:
    if ext in {".csv", ".xlsx", ".xls", ".json"}:
        return True
    download_types = [
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
        "text/csv",
        "application/json",
    ]
    return any(t in content_type for t in download_types)


def _extract_pairs_from_file(path: Path) -> List[dict]:
    ext = path.suffix.lower()
    if ext == ".csv":
        return _extract_pairs_from_csv(path)
    if ext in {".xlsx", ".xls"}:
        return _extract_pairs_from_xlsx(path)
    if ext == ".json":
        return _extract_pairs_from_json(path)
    return []


def _extract_pairs_from_csv(path: Path) -> List[dict]:
    out = []
    try:
        with path.open("r", newline="", encoding="utf-8-sig") as handle:
            reader = csv.reader(handle)
            for row in reader:
                out.extend(_scan_cells_for_pairs(row, source=str(path)))
    except Exception:
        return []
    return out


def _extract_pairs_from_xlsx(path: Path) -> List[dict]:
    out = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return out

    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                values = ["" if v is None else str(v) for v in row]
                out.extend(_scan_cells_for_pairs(values, source=f"{path}:{ws.title}"))
    finally:
        wb.close()
    return out


def _extract_pairs_from_json(path: Path) -> List[dict]:
    out = []
    try:
        data = json.loads(path.read_text(encoding="utf-8", errors="ignore"))
    except Exception:
        return out

    def walk(obj):
        if isinstance(obj, dict):
            values = [str(v) for v in obj.values() if isinstance(v, (str, int, float))]
            out.extend(_scan_cells_for_pairs(values, source=str(path)))
            for v in obj.values():
                walk(v)
        elif isinstance(obj, list):
            for item in obj:
                walk(item)

    walk(data)
    return out


def _extract_pairs_from_text(text: str, source: str) -> List[dict]:
    plain = _strip_html(text)
    lines = [line.strip() for line in plain.splitlines() if line.strip()]
    out = []
    for line in lines:
        out.extend(_scan_cells_for_pairs([line], source=source))
    return out


def _scan_cells_for_pairs(cells: Iterable[str], source: str) -> List[dict]:
    joined = " | ".join(cells)
    normalized_joined = normalize_text(joined)
    out = []

    # Look for 8/6/4-digit HSN-like numbers and nearby descriptions.
    for match in re.finditer(r"\b\d{4,8}\b", joined):
        code = normalize_hsn_digits(match.group(0))
        if len(code) not in {4, 6, 8}:
            continue

        start = max(0, match.start() - 80)
        end = min(len(joined), match.end() + 120)
        context = joined[start:end]
        context = re.sub(r"\s+", " ", context).strip()

        description = _clean_description_fragment(context, code)
        if not description:
            description = normalized_joined[:120]

        out.append(
            {
                "hsn_code": code,
                "description": description,
                "source": source,
            }
        )
    return out


def _clean_description_fragment(fragment: str, code: str) -> str:
    text = fragment.replace(code, " ")
    text = re.sub(r"\bhsn\b|\bsac\b|\bcode\b", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"[^a-zA-Z0-9\s,\-]", " ", text)
    text = re.sub(r"\s+", " ", text).strip(" -,:")
    return text[:160]


def _extract_links(base_url: str, html_text: str) -> List[str]:
    links = []
    for href in re.findall(r"href\s*=\s*['\"]([^'\"]+)['\"]", html_text, flags=re.IGNORECASE):
        href = html.unescape(href).strip()
        if not href or href.startswith("javascript:"):
            continue
        full = urllib.parse.urljoin(base_url, href)
        if full.startswith("http://") or full.startswith("https://"):
            links.append(full)
    return links


def _normalize_url(url: str) -> str:
    parsed = urllib.parse.urlparse(url)
    path = parsed.path or "/"
    if path != "/" and path.endswith("/"):
        path = path[:-1]

    # Keep query only if likely relevant to classification content.
    query = parsed.query
    query_l = query.lower()
    if query and not any(token in query_l for token in RELEVANT_TOKENS):
        query = ""

    normalized = urllib.parse.urlunparse((parsed.scheme, parsed.netloc, path, "", query, ""))
    return normalized


def _is_relevant_link(url: str) -> bool:
    parsed = urllib.parse.urlparse(url)
    ext = Path(parsed.path).suffix.lower()
    if ext in ASSET_EXTENSIONS:
        return False

    joined = f"{parsed.path} {parsed.query}".lower()
    if any(token in joined for token in EXCLUDE_TOKENS):
        return False

    # Always allow downloadable data files.
    if ext in {".csv", ".xlsx", ".xls", ".json", ".xml", ".zip"}:
        return True

    # Keep only potentially relevant content pages.
    return any(token in joined for token in RELEVANT_TOKENS)


def _allow_domain(seed_urls: List[str], candidate: str) -> bool:
    candidate_domain = urllib.parse.urlparse(candidate).netloc.lower()
    if not candidate_domain:
        return False

    seed_domains = {urllib.parse.urlparse(url).netloc.lower() for url in seed_urls}
    # Allow same domains and common subdomains.
    for domain in seed_domains:
        if candidate_domain == domain or candidate_domain.endswith("." + domain):
            return True
    return False


def _strip_html(text: str) -> str:
    text = re.sub(r"<script[\s\S]*?</script>", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"<style[\s\S]*?</style>", " ", text, flags=re.IGNORECASE)
    text = re.sub(r"<[^>]+>", " ", text)
    text = html.unescape(text)
    text = re.sub(r"\s+", " ", text)
    return text


def _safe_name_from_url(url: str) -> str:
    parsed = urllib.parse.urlparse(url)
    base = Path(parsed.path).name or "source"
    keep = "abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789._-"
    cleaned = "".join(ch if ch in keep else "_" for ch in base)
    if not cleaned:
        cleaned = "source_file"
    return cleaned[:120]


def _dedup_pairs(rows: List[dict]) -> List[dict]:
    seen: Set[Tuple[str, str]] = set()
    out = []
    for row in rows:
        code = row.get("hsn_code", "")
        desc = normalize_text(row.get("description", ""))
        if not code or not desc:
            continue
        key = (code, desc)
        if key in seen:
            continue
        seen.add(key)
        out.append(
            {
                "hsn_code": code,
                "description": row.get("description", ""),
                "source": row.get("source", ""),
            }
        )
    return out


def _write_pairs_csv(path: Path, rows: List[dict]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["hsn_code", "description", "source"])
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
