from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import List

from openpyxl import Workbook

from gst_hsn_tool.config import LEARNING_DB_PATH, LEARNING_SAVE_THRESHOLD, PUBLISH_MATCH_THRESHOLD
from gst_hsn_tool.learning import LearningMemory
from gst_hsn_tool.loader import load_client_data, load_hsn_master
from gst_hsn_tool.matcher import HsnMatcher, select_primary, status_from_score
from gst_hsn_tool.models import MatchCandidate


def run_pipeline(client_path: Path, hsn_master_path: Path, output_path: Path) -> dict:
    client_rows = load_client_data(client_path)
    hsn_rows = load_hsn_master(hsn_master_path)
    matcher = HsnMatcher(hsn_rows)
    learning = LearningMemory(Path(LEARNING_DB_PATH))

    if output_path.suffix.lower() == ".xlsx" and len(client_rows) > 1_000_000:
        raise ValueError(
            "Excel output has row limits. For 20 lakh rows, choose CSV output path (for example output.csv)."
        )

    mapped_rows = []
    review_rows = []
    audit_rows = []
    cache = {}
    learned_exact_hits = 0
    learned_fuzzy_hits = 0
    to_learn = []

    for idx, row in enumerate(client_rows):
        cache_key = (
            row["description_norm"],
            row["category_norm"],
            row["client_hsn_norm"],
        )
        if cache_key in cache:
            primary, reason, candidates = cache[cache_key]
        else:
            learned = learning.lookup_exact(
                description_norm=row["description_norm"],
                category_norm=row["category_norm"],
                client_hsn_norm=row["client_hsn_norm"],
            )
            if learned is not None:
                learned_exact_hits += 1
                primary = MatchCandidate(
                    hsn8=learned.resolved_hsn8,
                    description=learned.resolved_description,
                    category=learned.resolved_category,
                    rate=learned.resolved_rate,
                    match_type="learned_exact",
                    score=95,
                    reason="Learned memory exact hit",
                )
                reason = "Learned memory exact hit"
                candidates = [primary]
            else:
                learned_fuzzy = learning.lookup_fuzzy(
                    description_norm=row["description_norm"],
                    category_norm=row["category_norm"],
                )
                if learned_fuzzy is not None:
                    learned_fuzzy_hits += 1
                    primary = MatchCandidate(
                        hsn8=learned_fuzzy.resolved_hsn8,
                        description=learned_fuzzy.resolved_description,
                        category=learned_fuzzy.resolved_category,
                        rate=learned_fuzzy.resolved_rate,
                        match_type="learned_fuzzy",
                        score=86,
                        reason="Learned memory fuzzy hit",
                    )
                    reason = "Learned memory fuzzy hit"
                    candidates = [primary]
                else:
                    candidates = matcher.resolve(
                        description_norm=row["description_norm"],
                        category_norm=row["category_norm"],
                        client_hsn_norm=row["client_hsn_norm"],
                    )
                    primary, reason = select_primary(candidates)
            cache[cache_key] = (primary, reason, candidates)

        candidate_codes = "|".join([c.hsn8 for c in candidates])
        candidate_scores = "|".join([str(c.score) for c in candidates])

        if primary is None:
            resolved_hsn = ""
            score = 0.0
            status = "manual_review"
            match_type = "none"
            rate = ""
        else:
            resolved_hsn = primary.hsn8
            score = primary.score
            status = status_from_score(score)
            match_type = primary.match_type
            rate = primary.rate

        suggested_hsn4 = resolved_hsn[:4] if resolved_hsn else ""
        suggested_hsn6 = resolved_hsn[:6] if resolved_hsn else ""
        suggested_hsn_description = primary.description if primary else ""
        suggested_category = primary.category if primary else row.get("category", "")

        if score >= PUBLISH_MATCH_THRESHOLD:
            published_category = suggested_category
            published_hsn4 = suggested_hsn4
            published_hsn6 = suggested_hsn6
            published_hsn_description = suggested_hsn_description
        else:
            # Keep primary columns blank for low-confidence guesses to avoid misleading filing output.
            published_category = ""
            published_hsn4 = ""
            published_hsn6 = ""
            published_hsn_description = ""

        mapped_row = {
            "product_name": row.get("description", ""),
            "category": published_category,
            "hsn4": published_hsn4,
            "hsn6": published_hsn6,
            "hsn_description": published_hsn_description,
            "suggested_category": suggested_category,
            "suggested_hsn4": suggested_hsn4,
            "suggested_hsn6": suggested_hsn6,
            "suggested_hsn_description": suggested_hsn_description,
            "row_index": idx,
            "product_id": row.get("product_id", ""),
            "description": row.get("description", ""),
            "client_hsn": row.get("client_hsn", ""),
            "resolved_hsn8": resolved_hsn,
            "resolved_rate": rate,
            "score": score,
            "status": status,
            "match_type": match_type,
            "top_candidates": candidate_codes,
            "candidate_scores": candidate_scores,
            "reason": reason,
        }
        mapped_rows.append(mapped_row)

        if resolved_hsn and score >= LEARNING_SAVE_THRESHOLD:
            to_learn.append(
                {
                    "input_description_norm": row["description_norm"],
                    "input_category_norm": row["category_norm"],
                    "input_client_hsn_norm": row["client_hsn_norm"],
                    "resolved_hsn8": resolved_hsn,
                    "resolved_description": suggested_hsn_description,
                    "resolved_category": suggested_category,
                    "resolved_rate": rate,
                    "match_type": match_type,
                    "score": score,
                }
            )

        if status != "auto_approved":
            review_rows.append(mapped_row)

        audit_rows.append(
            {
                "row_index": idx,
                "input_description_norm": row["description_norm"],
                "input_category_norm": row["category_norm"],
                "input_client_hsn_norm": row["client_hsn_norm"],
                "selected_hsn8": resolved_hsn,
                "score": score,
                "status": status,
                "reason": reason,
                "candidates": candidate_codes,
            }
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.suffix.lower() == ".csv":
        _write_csv(output_path, mapped_rows)
        final_output_path = output_path
    else:
        wb = Workbook()
        ws_mapped = wb.active
        ws_mapped.title = "mapped"
        _write_sheet(ws_mapped, mapped_rows)

        ws_review = wb.create_sheet("review_queue")
        _write_sheet(ws_review, review_rows)

        ws_audit = wb.create_sheet("audit_log")
        _write_sheet(ws_audit, audit_rows)
    final_output_path = _save_workbook_with_fallback(wb, output_path)

    learned_saved = learning.upsert_many(to_learn)

    return {
        "total_rows": len(mapped_rows),
        "review_rows": len(review_rows),
        "auto_approved_rows": len(mapped_rows) - len(review_rows),
        "unique_products": len(cache),
        "learned_exact_hits": learned_exact_hits,
        "learned_fuzzy_hits": learned_fuzzy_hits,
        "learned_saved": learned_saved,
        "output_path": str(final_output_path),
    }


def _write_sheet(worksheet, rows: List[dict]) -> None:
    if not rows:
        worksheet.append(["no_data"])
        return

    headers = list(rows[0].keys())
    worksheet.append(headers)
    for row in rows:
        worksheet.append([row.get(col, "") for col in headers])


def _write_csv(path: Path, rows: List[dict]) -> None:
    import csv

    if not rows:
        path.write_text("no_data\n", encoding="utf-8")
        return

    headers = [
        "product_name",
        "category",
        "hsn4",
        "hsn6",
        "hsn_description",
        "suggested_category",
        "suggested_hsn4",
        "suggested_hsn6",
        "suggested_hsn_description",
        "resolved_hsn8",
        "score",
        "status",
        "reason",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in headers})


def _save_workbook_with_fallback(workbook: Workbook, preferred_path: Path) -> Path:
    try:
        workbook.save(preferred_path)
        return preferred_path
    except PermissionError:
        # File is likely open in Excel; save to a new timestamped name instead.
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        fallback = preferred_path.with_name(f"{preferred_path.stem}_{stamp}{preferred_path.suffix}")
        workbook.save(fallback)
        return fallback
