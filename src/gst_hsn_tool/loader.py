from __future__ import annotations

import csv
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook
import xlrd

from gst_hsn_tool.config import REQUIRED_CLIENT_COLUMNS, REQUIRED_HSN_COLUMNS
from gst_hsn_tool.utils import normalize_hsn_digits, normalize_text


CLIENT_DESCRIPTION_ALIASES = [
    "description",
    "product_name",
    "product",
    "item",
    "item_name",
    "name",
]


def _read_csv(path: Path) -> List[Dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        return [dict(row) for row in reader]


def _read_excel(path: Path) -> List[Dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = []
    for row in rows[1:]:
        item = {}
        for i, header in enumerate(headers):
            if not header:
                continue
            value = row[i] if i < len(row) else ""
            item[header] = "" if value is None else str(value)
        data.append(item)
    return data


def _read_xls(path: Path) -> List[Dict[str, str]]:
    book = xlrd.open_workbook(path)
    sheet = book.sheet_by_index(0)
    if sheet.nrows == 0:
        return []

    headers = [str(sheet.cell_value(0, c)).strip() for c in range(sheet.ncols)]
    out = []
    for r in range(1, sheet.nrows):
        row_obj = {}
        for c, header in enumerate(headers):
            if not header:
                continue
            value = sheet.cell_value(r, c)
            row_obj[header] = "" if value is None else str(value)
        out.append(row_obj)
    return out


def _read_table(path: Path) -> List[Dict[str, str]]:
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return _read_excel(path)
    if suffix == ".xls":
        return _read_xls(path)
    if suffix == ".csv":
        return _read_csv(path)
    raise ValueError(f"Unsupported file type: {path}")


def load_client_data(path: Path) -> List[Dict[str, str]]:
    rows = _read_table(path)
    if not rows:
        raise ValueError("Client input file has no rows.")

    columns = set(rows[0].keys())
    if "description" not in columns:
        alias = next((c for c in CLIENT_DESCRIPTION_ALIASES if c in columns), None)
        if alias:
            for row in rows:
                row["description"] = row.get(alias, "")
        elif len(columns) == 1:
            sole = next(iter(columns))
            rebuilt = []

            # Handles files where users provide only a single product-name column.
            if normalize_text(sole) not in CLIENT_DESCRIPTION_ALIASES and sole.strip():
                rebuilt.append({"description": sole})

            for row in rows:
                value = row.get(sole, "")
                if str(value).strip():
                    rebuilt.append({"description": value})
            rows = rebuilt
        else:
            missing = [col for col in REQUIRED_CLIENT_COLUMNS if col not in columns]
            raise ValueError(f"Missing required client columns: {missing}")

    out = []
    for row in rows:
        prepared = dict(row)
        for optional_col in ["product_id", "category", "client_hsn", "gst_rate"]:
            prepared.setdefault(optional_col, "")

        prepared["description_norm"] = normalize_text(prepared.get("description", ""))
        prepared["category_norm"] = normalize_text(prepared.get("category", ""))
        prepared["client_hsn_norm"] = normalize_hsn_digits(prepared.get("client_hsn", ""))
        out.append(prepared)
    return out


def load_hsn_master(path: Path) -> List[Dict[str, str]]:
    rows = _read_table(path)
    if not rows:
        raise ValueError("HSN master file has no rows.")

    columns = set(rows[0].keys())
    missing = [col for col in REQUIRED_HSN_COLUMNS if col not in columns]
    if missing:
        raise ValueError(f"Missing required HSN columns: {missing}")

    out = []
    for row in rows:
        prepared = dict(row)
        for optional_col in ["category", "rate", "aliases"]:
            prepared.setdefault(optional_col, "")

        hsn8 = normalize_hsn_digits(prepared.get("hsn8", ""))
        if len(hsn8) != 8:
            continue

        prepared["hsn8"] = hsn8
        prepared["description_norm"] = normalize_text(prepared.get("description", ""))
        prepared["category_norm"] = normalize_text(prepared.get("category", ""))
        prepared["aliases_norm"] = [
            normalize_text(piece)
            for piece in str(prepared.get("aliases", "")).split("|")
            if normalize_text(piece)
        ]
        prepared["hsn4"] = hsn8[:4]
        prepared["hsn6"] = hsn8[:6]
        out.append(prepared)

    if not out:
        raise ValueError("No valid 8-digit HSN rows found in master file.")
    return out
